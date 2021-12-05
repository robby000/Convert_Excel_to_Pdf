# This is a program to create pdf invoices from data in an excel spreadsheet.


#  The openpyxl, reportlab and pillow(PIL) libraries must be installed


############################################################################################

# Imports

import openpyxl
from os import getcwd
from datetime import date
from time import sleep
from PIL import Image
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import Table
from reportlab.lib import colors

#  import fonts
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf'))
pdfmetrics.registerFont(TTFont('VeraBd', 'VeraBd.ttf'))


#############################################################################################
# CONSTANTS

#  Page information
page_height = 3508
page_width = 2480

# constant margin on left
margin_left = 250

#  Company Information

# dictionary for header information
header_info = {
    "company_name": "Company Name",
    "company_address1": "House Number and Street",
    "company_address2": "City",
    "company_address3": "Postcode",
    "email": "Email",
    "phone_number": "Phone Number",
}

# dictionary for footer information
footer_info = {
    "Therapist": "Therapist's Name",
    "Account Number": "Account Number: *******",
    "Sort Code": "Sort Code: **-**-**",
    "company_registration_num": "Company Registration No: *********"
}


###############################################################################################
#  define functions

#  function to create invoice header
def invoice_header():
    #  insert image with coordinates and size
    c.drawInlineImage(im, x=1780, y=2828, width=450, height=450, )

    #  insert invoice heading
    c.setFont('VeraBd', 80)
    invoice = 'INVOICE: ' + str(invoice_num)
    margin_bottom = 3200  # set margin bottom variable
    c.drawString(margin_left, margin_bottom, invoice)
    margin_bottom = margin_bottom - 135

    # Invoice on
    c.setFont('VeraBd', 50)
    c.drawString(1250, margin_bottom, "Issued on:")

    # Date
    today = date.today()
    c.setFont('Vera', 50)
    d1 = today.strftime("%d/%m/%Y")
    data_header = d1
    c.drawString(1250, margin_bottom - 70, data_header)

    #  issued by
    c.setFont('VeraBd', 50)
    c.drawString(margin_left, margin_bottom, 'Issued by:')
    margin_bottom = margin_bottom - 70

    #  Insert company information from list
    for info in header_info.values():
        c.setFont('Vera', 50)
        header_line = info
        c.drawString(margin_left, margin_bottom, header_line)
        margin_bottom = margin_bottom - 70

    margin_bottom = margin_bottom - 40  # create space for next part

    # Therapist
    c.setFont('VeraBd', 50)
    c.drawString(1780, margin_bottom, "Therapist: ")

    # Therapist Name
    c.setFont('Vera', 50)
    therapist_name = footer_info['Therapist']
    c.drawString(1780, margin_bottom - 80, therapist_name)

    # invoice to
    c.setFont('VeraBd', 50)
    c.drawString(margin_left, margin_bottom, "Issued to:")
    margin_bottom = margin_bottom - 70

    # School or Private
    # if it is private print clients name
    if str(sheet.title) == "Private":
        client = sheet.cell(row=2, column=1).value  # Retrieve client name from excel file
        c.setFont('Vera', 50)
        client_name = "Parents of: " + client
        c.drawString(margin_left, margin_bottom, client_name)

    # otherwise if its a school print name and address
    else:
        c.setFont('Vera', 50)
        c.drawString(margin_left, margin_bottom, school_name)

        # Using school name to retrieve school address from dictionary
        c.drawString(margin_left, margin_bottom - 70, school_address[school_name])


#  function to create invoice table
def invoice_table():
    f = Table(rows_list, colWidths=600, rowHeights=110)
    t_style = [
        ('GRID', (0, 0), (-1, -1), 5, colors.darkblue),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),  # First row backgroud colour
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'VeraBd'),  # First row font
        ('FONTNAME', (0, 1), (-1, -1), 'Vera'),
        ('FONTSIZE', (0, 0), (-1, 0), 55),  # First row font size
        ('FONTSIZE', (0, 1), (-1, -1), 50),
    ]

    # Because the x and y of the pdf are set from the bottom I need to change the coordination relative to the
    # amount of sessions
    table_margin_bottom = 0  # create bottom margin variable

    if number <= 6:
        table_margin_bottom = 1750
    elif 6 < number <= 11:
        table_margin_bottom = 1200
    elif 11 < number <= 16:
        table_margin_bottom = 700

    f.setStyle(t_style)
    f.wrap(page_width, page_height)
    f.drawOn(c, margin_left + 90, table_margin_bottom)
    margin_bottom = table_margin_bottom - 100

    # invoice total if this is the only page of pdf
    if number_page_2 is None:
        c.setFont('VeraBd', 50)
        total_end = "Total: £" + str(total)
        c.drawString(1810, margin_bottom, total_end)


#  function to create invoice footer
def invoice_footer():
    footer_margin_bottom = 370

    c.setFont('Vera', 45)
    c.drawString(margin_left + 80, footer_margin_bottom, footer_info['Account Number'])

    c.drawString(1685, footer_margin_bottom, footer_info['Sort Code'])
    footer_margin_bottom = footer_margin_bottom - 100

    c.drawString(870, footer_margin_bottom, footer_info['company_registration_num'])
    footer_margin_bottom = footer_margin_bottom - 110

    if number_page_2 is None:
        page = "Page " + str(page_number) + " of 1"
    else:
        page = "Page " + str(page_number) + " of 2"
    c.drawString(1110, footer_margin_bottom, page)


#####################################################################################################
# Importing data from external files

#  get name of excel file from user
file_name = input("This program converts data from excel files into pdf invoices. What is the name of the excel file? ")

#  get location of excel file from user
while True:
    location_input = input("Is your excel file located in the same directory as this program? ")
    if location_input == "yes":
        location = getcwd()  # Create relative path for excel file
        break
    # if not in same directory ask user for file location
    elif location_input == "no":
        location = input("Please enter the file's location? ")
        break
    else:
        print("Please answer yes or no!")

excel = location + "\\" + file_name + ".xlsx"

#  import excel file and deal with errors
while True:
    try:
        wb = openpyxl.load_workbook(excel)
        break  # break from continuous loop if there is no error
    except PermissionError:  # This is for the common error when the excel file was left open
        print("Please close the excel file.")
        sleep(5)  # Give the user a chance close the file
    except FileNotFoundError:  # Wrong filename error
        print("ERROR: File not found. Please recheck the excel file's name and location, then restart program.")
        exit(1)

#  import logo image
im = Image.open('logo.png')

# Open file and retrieve invoice number
invoice_num_file = open("Invoice Number", "r")
invoice_num = int(invoice_num_file.read())
invoice_num_file.close()

#  create dictionary of school addresses from text file
school_address = {}
with open("address.txt") as conf:
    for line in conf:
        if ":" in line:
            name, value = line.split(":")
            school_address[name] = str(value)
for key, value in school_address.items():  # removing line break at the end of dictionary value
    school_address[key] = value.rstrip()

##################################################################################################

#  Running the program

for sheet in wb:  # creating a new invoice for each sheet in the excel file

    # Extracting school name from sheet title
    school_name = str(sheet.title)

    # Finding out number of sessions(rows)
    number = 2  # Reset number for new invoice
    for a in range(2, 100):
        if sheet.cell(row=a, column=2).value is not None:
            number = number + 1

    #  checking if number of sessions are too large for one page or even too large for two (which is not likely)
    page_number = 1  # Reset page number
    number_page_2 = None
    if number > 16:
        number_page_2 = number
        number = 16  # for first page
    elif number > 30:
        print("The data in the sheet for " + school_name + " is too big for this invoice please consult the program's "
                                                           "author")
    else:
        pass  # because it is less than 16 so can fit on one page

    # Creating list of lists from the values in the excel file by row to be used in the table
    rows_list = []
    rows = sheet.iter_rows(min_col=1, max_row=number - 1, values_only=True)
    # get the values and add to list of lists
    for row in rows:
        rows_list.append(row)

    #   Create list of integers from session prices in excel file then use addition to get total
    price_list = []  # Reset price list
    for k in range(2, number):
        price = sheet.cell(row=k, column=3).value
        price_list.append(price)

    total = sum(price_list)

    #  Set pfd document name and size
    c = canvas.Canvas("Compay Name Invoice " + str(invoice_num) + ".pdf")
    c.setPageSize((page_width, page_height))

    #  creating invoice header from function
    invoice_header()

    # create table from function
    invoice_table()

    # create invoice footer from function
    invoice_footer()

    # creating first page of the pdf
    c.showPage()

    # create extra page in the PDF if there is too much data for the one page
    if number_page_2 is not None:

        page_number = 2

        # Creating list of lists from remaining values in the excel file by row to be used in the table
        rows_list = [('Client Name', 'Date', 'Price per session')]
        rows = sheet.iter_rows(min_col=1, min_row=number, max_row=number_page_2 - 1, values_only=True)

        for row in rows:
            rows_list.append(row)

        # getting total for sessions of both pages
        for k in range(2, number_page_2):
            price = sheet.cell(row=k, column=3).value
            price_list.append(price)

        invoice_header()

        invoice_table()

        # printing total for sessions of both pages
        c.setFont('VeraBd', 50)
        text = "Total: £" + str(total)
        c.drawString(1810, 550, text)

        invoice_footer()

    #  update invoice number for next excel sheet or next time
    invoice_num = invoice_num + 1

    c.save()

#####################################################################################################


#  save new invoice number for next time
invoice_num_file = open("Invoice Number", "r+")
invoice_num_file.write(str(invoice_num))
invoice_num_file.close()
